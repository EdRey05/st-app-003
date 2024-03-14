'''
App made by:
    Eduardo Reyes Alvarez, Ph.D.
Contact:
    eduardo_reyes09@hotmail.com

App version: 
    V11 (Mar 14, 2024): Some more exceptions handled. When the clinical and RNA patient IDs do not
                        match, and when there are NaNs in the time to event column. The logging
                        has not been revised yet. 
'''
###################################################################################################

# Import required libraries

import io
import logging
from typing import List
from collections import OrderedDict

import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import altair as alt 

import streamlit as st
from streamlit_searchbox import st_searchbox

from lifelines import KaplanMeierFitter
from lifelines.plotting import add_at_risk_counts

###################################################################################################

# App configuration and layout
st.set_page_config(
    page_title="Tool 003 - App by Eduardo",
    page_icon=":chart_with_downwards_trend:",
    layout="wide",
    initial_sidebar_state="expanded")

# Use these colors for altair charts
st.session_state["alt_colors"] = ["#76448A", "#B03A2E", "#1E8449", "#1F618D", "#34495E ",  
                                "#D68910", "#707B7C", "#E67E22", "#2E86C1", "#E74C3C",
                                "#2C3E50", "#F1C40F", "#3498DB", "#D35400", "#27AE60"]
# Title
st.title("Interactive Kaplan-Meier plot generator")
st.markdown('<hr style="margin-top: +2px; margin-bottom: +2px; border-width: 5px;">', unsafe_allow_html=True)

# Sidebar - Initial widgets
with st.sidebar:
    uploaded_files = st.file_uploader(label="Upload a clinical file (and optionally, a RNA file)", 
                                    type=["txt"], accept_multiple_files=True)
    st.markdown('<hr style="margin-top: 1px; margin-bottom: 1px; border-width: 5px;">', unsafe_allow_html=True)
    start_button = st.button(label="Begin", type="secondary")
    restart_button = st.button(label="Start over", type="secondary")

###################################################################################################

# Function to setup the logging configuration
def logging_setup():
    
    # Check if logging has already been initialized
    if "log_created" not in st.session_state:
        
        # Configure the logging settings
        logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

        # Create a logger
        logger = logging.getLogger()

        # Clear the existing log file or create a new empty one
        open("MyLog.txt", "w").close()

        # Create a file handler
        file_handler = logging.FileHandler("MyLog.txt")
        file_handler.setLevel(logging.INFO)

        # Create a formatter and add it to the file handler
        formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        file_handler.setFormatter(formatter)

        # Add the file handler to the logger
        logger.addHandler(file_handler)

        # Log an initial message
        logger.info(f"Log file created or cleared. \n")

        # Update the run variable so this set up process gets executed only once
        st.session_state["logger"] = logger
        st.session_state["log_created"] = True

###################################################################################################

# Function to prompt the user to upload the files if not already uploaded
def load_input_files(uploaded_files):

    # Check if one or two files were uploaded
    clinical_file = next((file for file in uploaded_files if file.name == "clinical.txt"), None)
    RNA_file = next((file for file in uploaded_files if file.name == "RNA.txt"), None)
    
    # Load the files available
    if clinical_file is not None:
        
        # Initialize the logging file 
        logging_setup()
        logger = st.session_state["logger"]

        # Load, log and save to the session state the mandatory file
        df_clinical = pd.read_csv(clinical_file, sep="\t", comment="#")
        logger.info(f"File found: clinical.txt \n")
        st.session_state["df_clinical"] = df_clinical

        # Load, log and save to the session state the optional file if it exists
        if RNA_file is not None:
            df_RNA = pd.read_csv(RNA_file, sep="\t")
            logger.info(f"File found: RNA.txt \n")
            st.session_state["df_RNA"] = df_RNA            
    else:
        st.sidebar.error("A clinical file is required!")
        st.stop()

###################################################################################################

# This function searches OS/PFS/RFS/DFS _STATUS and _MONTHS columns in the clinical data
# This function transposes the RNA dataset to have gene names as columns and patient IDs as rows (like the clinical data)

def file_preprocessing():

    # Get the df(s) and the logger
    df_clinical = st.session_state.get("df_clinical")
    df_RNA = st.session_state.get("df_RNA", None)
    logger = st.session_state.get("logger")

    ################### Processing for the clinical dataframe ##################

    # Log the original dataframe
    logger.info(f"Preview of the original clinical dataset: \n {df_clinical.iloc[:15, :10].to_string()} \n")
    logger.info(f"Data types of columns in the original clinical dataset: \n {df_clinical.dtypes.to_string()} \n\n")

    # Prepare the variable with the reordered column names
    clinical_columns_main = ["PATIENT_ID"]
    time_to_event_options = []
    event_observation_options = []

    # Search for possible metric measurements (Overall/Recurrence-Free/Progression-Free/Disease-Free/Disease-specific Survival
    for metric in ["OS", "RFS", "PFS", "DFS", "DSS"]:
        if metric+"_MONTHS" in df_clinical.columns and metric+"_STATUS" in df_clinical.columns:
            clinical_columns_main.append(metric+"_MONTHS")
            time_to_event_options.append(metric+"_MONTHS")
            clinical_columns_main.append(metric+"_STATUS")
            event_observation_options.append(metric+"_STATUS")

    # Search for a column of Vital status or Cause of Death (this is optional and may provide useful information)
    for extra_metric in ["VITAL_STATUS", "CAUSE_OF_DEATH"]:
        if extra_metric in df_clinical.columns:
            clinical_columns_main.append(extra_metric)
            event_observation_options.append(extra_metric)

    # Order alphabetically the remaining columns
    clinical_columns_extra = [col for col in df_clinical.columns if col not in clinical_columns_main]
    clinical_columns_extra.sort()

    # Apply the re-ordering to the df
    clinical_columns_ordered = clinical_columns_main + clinical_columns_extra
    df_clinical = df_clinical[clinical_columns_ordered] 

    # Log the re-arranged dataframe
    logger.info(f"Preview of the pre-processed clinical dataset: \n {df_clinical.iloc[:15, :10].to_string()} \n")
    logger.info(f"Data types of columns in the pre-processed clinical dataset: \n {df_clinical.dtypes.to_string()} \n\n")

    ################### Processing for the RNA dataframe ###################

    # If an RNA file was uploaded, then the df is not empty
    if df_RNA is not None:
        # Log the original dataframe
        logger.info(f"Preview of the original RNA dataset: \n {df_RNA.iloc[:15, :10].to_string()} \n")
        logger.info(f"Data types of some columns in the original RNA dataset: \n {df_RNA.iloc[:, :10].dtypes.to_string()} \n\n")
        
        # Drop the "Entrez_Gene_Id" column if exists
        if "Entrez_Gene_Id" in df_RNA.columns:
            df_RNA.drop("Entrez_Gene_Id", axis=1, inplace=True)
        
        # Rename the "Hugo_Symbol" column to "PATIENT_ID" as it appears in the clinical df
        df_RNA.rename(columns={"Hugo_Symbol": "PATIENT_ID"}, inplace=True)
        
        # Transpose the dataframe, making the content of the "PATIENT_ID" column the new column names
        df_RNA = df_RNA.set_index("PATIENT_ID").T
        
        # Sort the gene names alphabetically
        df_RNA.sort_index(axis=1, inplace=True)
        
        # Reset the index to a numerical index
        df_RNA = df_RNA.reset_index()
        
        # Rename the "index" column to "PATIENT_ID"
        df_RNA.rename(columns={"index": "PATIENT_ID"}, inplace=True)
        
        # Sort Patient IDs and reset the index
        df_RNA = df_RNA.sort_values("PATIENT_ID").reset_index(drop=True)

        # Log the re-arranged dataframe
        logger.info(f"Preview of the pre-processed RNA dataset: \n {df_RNA.iloc[:15, :10].to_string()} \n")
        logger.info(f"Data types of some columns in the pre-processed RNA dataset: \n {df_RNA.iloc[:, :10].dtypes.to_string()} \n\n")

    ###################

    # Save the neccesary data for the next steps in the session state
    st.session_state.update({"df_clinical": df_clinical,
                            "df_RNA": df_RNA,
                            "time_to_event_options": time_to_event_options,
                            "event_observation_options": event_observation_options})

    # Also add the options of genes if a RNA file was uploaded
    if df_RNA is not None:
        st.session_state["gene_list"] = tuple(df_RNA.columns[1:].tolist())

###################################################################################################

# Main function to prepare and display the interactive widgets and subwidgets
def widget_preparation():

    # Get required variables from the session state
    time_to_event_options = st.session_state.get("time_to_event_options")
    event_observation_options = st.session_state.get("event_observation_options")
    logger = st.session_state.get("logger")

    logger.info(f"---------------User interaction with the widgets starts here--------------- \n")

    # Create the layout for the widgets
    col_1_row_1, col_2_row_1 = st.columns(2, gap="medium")
    col_1_row_2, col_2_row_2 = st.columns(2, gap="medium")
    st.markdown('<hr style="margin-top: -15px; margin-bottom: -15px;">', unsafe_allow_html=True)
    col_1_row_3, col_2_row_3, col_3_row_3, col_4_row_3 = st.columns([2, 2, 1.35, 1.15], gap="medium")
    st.markdown('<hr style="margin-top: +10px; margin-bottom: +10px; border-width: 5px;">', unsafe_allow_html=True)
    col_1_row_14, col_2_row_14, col_3_row_14 = st.columns([0.5, 9, 0.5], gap="medium")
    col_1_row_15, col_2_row_15, col_3_row_15, col_4_row_15 = st.columns(4, gap="small")
    
    # Save the columns and containers in the session state
    widget_and_output_areas = [col_1_row_1, col_2_row_1,
                                col_1_row_2, col_2_row_2,
                                col_1_row_3, col_2_row_3, col_3_row_3, col_4_row_3,
                                col_1_row_14, col_2_row_14, col_3_row_14,
                                col_1_row_15, col_2_row_15, col_3_row_15, col_4_row_15]
    st.session_state["widget_and_output_areas"] = widget_and_output_areas
    
    # Time to event widget and callback function
    with col_1_row_1:
        time_to_event_dropdown = st.selectbox(label="Select the time-to-event column", 
                                options=["Click here to select..."] + time_to_event_options)
    if time_to_event_dropdown:
        time_to_event_dropdown_handler(time_to_event_dropdown)

    # Event observation widget and callback function
    with col_2_row_1:
        event_observation_dropdown = st.selectbox(label="Select the event observation column", 
                            options=["Click here to select..."] + event_observation_options)
    if event_observation_dropdown:
        event_observation_dropdown_handler(event_observation_dropdown)

    # Show widgets to generate+save the plot (with their callback functions), and to customize it
    with st.sidebar:
        generate_plot_button = st.button(label="Generate/Update plot", type="primary")
        st.markdown('<hr style="margin-top: 1px; margin-bottom: 1px; border-width: 5px;">', unsafe_allow_html=True)
        
        # Customize the plot
        st.write("#### Customize your plot here  ⬇️⬇️")
        CI_checkbox = st.checkbox(label="Show Confidence Intervals", value=True)
        move_labels_checkbox = st.checkbox(label="Move legend to the side", value=False)
        at_risk_checkbox = st.checkbox(label="Show at-risk table", value=False)
        sample_percent_slider = st.slider(label="Datapoints to plot (%)", min_value=50, max_value=100, value=95)
    
    # Add the plot customization options to the session state and save button to enable it later
    st.session_state["CI_checkbox"] = CI_checkbox
    st.session_state["move_labels_checkbox"] = move_labels_checkbox
    st.session_state["at_risk_checkbox"] = at_risk_checkbox
    st.session_state["sample_fraction"] = sample_percent_slider / 100
    
    # Control the rerun of the plot+excel generation process as not all interactions should trigger it
    if generate_plot_button or "logged_figure" in st.session_state:
        KM_plot_area = col_2_row_14
        
        # When reruns occur due to other widgets, show the same plot unless the user wants a new one
        previous_plot = st.session_state.get("logged_figure", False)
        if previous_plot and not generate_plot_button:
            with KM_plot_area:
                st.image(previous_plot)
        else:
            KM_figure = pass_KM_parameters()
            with KM_plot_area:
                KM_plot_area.empty()
                st.pyplot(KM_figure)

        # In any case the function to save the files should be executed (it handles logged files)
        figure_bytes, plot_filename, excel_bytes, excel_filename = save_KM_results(generate_plot_button)

        # Show the download buttons for the current data
        with col_2_row_15:
            download_plot = st.download_button(label="Download Plot", data=figure_bytes, 
                            file_name=plot_filename, type="primary", mime="image/png")
        with col_3_row_15:
            download_excel = st.download_button(label="Download Raw Data", data=excel_bytes, 
                            file_name=excel_filename, type="primary",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        # Log the download of any file
        if download_plot:
            logger.info(f"A KM plot with the name {plot_filename} has been downloaded \n")
        if download_excel:
            logger.info(f"An excel file with the name {excel_filename} has been downloaded \n")
    elif "subgroup_buttons_selection" in st.session_state and st.session_state["subgroup_buttons_selection"] == "Using variable(s)":
        # Show a message to tell the user where the plot will be shown (middle, below the variable repeats)
        with col_2_row_14:
            st.markdown('<div style="display: flex; justify-content: center;">'
                '<div style="background-color: #1eb53a; padding: 10px; width: fit-content;">'
                '<span style="font-weight: bold;">Your KM plot will be shown here!</span>'
                '</div></div>', unsafe_allow_html=True)

###################################################################################################

# Function to display the output of time_to_event_dropdown (histogram)
def time_to_event_dropdown_handler(change):
    
    # Save the current selection to the session state
    st.session_state["time_to_event_selection"] = change

    # Get the required variables from the session state
    df_clinical = st.session_state.get("df_clinical")
    logger = st.session_state.get("logger")
    time_to_event_output = st.session_state["widget_and_output_areas"][2]

    # Clear the output area and return early if the default option is selected back 
    if change == "Click here to select...":
        with time_to_event_output: 
            time_to_event_output.empty()
        return

    # If the selection is a column on the clinical dataframe, display a histogram
    column_name = change
    if column_name in df_clinical.columns:
        time_column = df_clinical[column_name].dropna()
        logger.info(f"The user selected: {column_name}     Widget: time_to_event_dropdown. \n")
        logger.info(f"Original dtype of {column_name}: {df_clinical[column_name].dtype}     Dtype once removing NANs: {time_column.dtype} \n")
        
        # Make histogram of values with altair and handle exceptions
        if time_column.dtype == "object":
            # First try to convert the column to a numeric type and if it fails we show a warning
            try:
                time_column = pd.to_numeric(time_column, errors="coerce")
                time_column = time_column.dropna()
            except:
                info_str = "Warning: Column type is not numeric."
                logger.warning("User attention required: The time to event column may not be numerical. \n")
                
                with time_to_event_output:
                    time_to_event_output.empty()
                    st.warning(info_str)
                st.stop()

        alt_data1 = pd.DataFrame({column_name: time_column})

        chart1 = alt.Chart(alt_data1).mark_bar(color="#BA4A00").encode(
                alt.X(column_name, type="quantitative", bin=alt.Bin(step=12)),
                alt.Y("count()", title="Patients"),
                ).properties(width=425, height=325
                ).configure_axis(labelColor="#3386BD")
        logger.info(f"A histogram was successfully made and displayed for: {column_name} \n")
            
    else:
        # If the column is not in the df, display an error and stop the app
        info_str = "Warning: Column not found in the dataframe."
        logger.error("User attention required: The time to event column name was not found in the df. \n")
        
        with time_to_event_output:
            time_to_event_output.empty()
            st.error(info_str)
        st.stop()

    # Clear the output area and display the histogram
    with time_to_event_output:
        time_to_event_output.empty()
        st.altair_chart(chart1)

###################################################################################################

# Function to display the output of event_observation_dropdown (bar chart)
def event_observation_dropdown_handler(change):

    # Save the current column selection to the session state
    st.session_state["event_observation_selection"] = change

    # Get the required variables from the session state
    df_clinical = st.session_state.get("df_clinical")
    logger = st.session_state.get("logger")
    event_observation_output_1 = st.session_state["widget_and_output_areas"][3]
    event_observation_output_2 = st.session_state["widget_and_output_areas"][4]
    event_observation_output_3 = st.session_state["widget_and_output_areas"][5]
    event_observation_output_4 = st.session_state["widget_and_output_areas"][6]
    alt_colors = st.session_state.get("alt_colors")

    # Clear the output and return early if the default option is selected back 
    if change == "Click here to select...":
        event_observation_output_1.empty()
        event_observation_output_2.empty()
        event_observation_output_3.empty()
        event_observation_output_4.empty()
        return
        
    # If the selection is a column on the clinical dataframe, display a bar chart
    column_name = change
    if column_name in df_clinical.columns:
        event_column = df_clinical[column_name]
        logger.info(f"The user selected: {column_name}     Widget: event_observation_dropdown. \n")
        logger.info(f"Dtype of {column_name}: {df_clinical[column_name].dtype}     Unique value counts: \n\t {event_column.value_counts(dropna=False).to_string()} \n")

        # Make a bar chart for unique values in the column and handle exceptions
        if event_column.dtype == "object":
            value_counts = event_column.value_counts(dropna=True)
            
            if df_clinical[column_name].nunique() > 10:
                logger.warning("User attention required: There may be something wrong with the event observation column as there are more than 15 unique values. \n")
            
            chart2 = alt.Chart(value_counts.reset_index()).mark_bar().encode(
                    alt.X(column_name, type="nominal", axis=alt.Axis(labelAngle=0)),
                    alt.Y("count", title="Patients"),
                    alt.Color(column_name, scale=alt.Scale(domain=list(value_counts.index), range=alt_colors))
                    ).properties(width=425, height=325
                    ).configure_axis(labelColor="#3386BD",
                    ).configure_legend(disable=True)
            logger.info(f"A bar chart was successfully made and displayed for: {column_name} \n")
        else:
            # If the column is not categorical, show a warning and stop the app
            info_str = "Warning: Column type is not categorical."
            logger.warning("User attention required: The event observation column may not be text-based. \n")
            
            with event_observation_output_1:
                event_observation_output_1.empty()
                st.warning(info_str)
            st.stop()
    else:
        # If the selection is not a column in the df, show an error and stop the app
        info_str = "Warning: Column not found in the dataframe."
        logger.error("User attention required: The event observation column name was not found in the df. \n")

        with event_observation_output_1:
            event_observation_output_1.empty()
            st.error(info_str)
        st.stop()

    # Clear the output area and display the bar chart
    with event_observation_output_1:
        event_observation_output_1.empty()
        st.altair_chart(chart2)
    
    # Get the unique values of the column currently selected    
    event_options = np.ndarray.tolist(df_clinical[change].unique())

    # Make subwidgets to specify the event to be observed so we can encode it in binary
    with event_observation_output_2:
        event_observation_output_2.empty()
        event_options_0 = st.multiselect(label="No event (0):", options=event_options)
    with event_observation_output_3:
        event_observation_output_3.empty()
        event_options_1 = st.multiselect(label="Event (1):", options=event_options)
    
    # Make a widget to ask the user if they want a curve for the whole dataset or divide it
    with event_observation_output_4:
        subgroup_buttons = st.radio(label="Make subgroups?", options=["None", "Using variable(s)"], index=0)
    if subgroup_buttons:
        subgroup_buttons_handler(subgroup_buttons)
    
    # Save the selected events in the session state
    st.session_state["event_0"] = event_options_0
    st.session_state["event_1"] = event_options_1

###################################################################################################

# Function to display the output of subgroup_buttons (slider)
def subgroup_buttons_handler(change):
    
    # Clear the plot only when changing between subgrouping options, persist the same otherwise
    old_selection = st.session_state.get("subgroup_buttons_selection", "None")
    if old_selection != change and "logged_figure" in st.session_state:
        del st.session_state["logged_figure"]

    # Save the subgroup selection to the session state 
    st.session_state["subgroup_buttons_selection"] = change

    # Get the required variables from the session state
    logger = st.session_state.get("logger")
    subgroup_buttons_output = st.session_state["widget_and_output_areas"][7]
    
    # If the user wants to make subgroups, ask the number of variables to use
    if change == "Using variable(s)":
        logger.info(f"The user selected: Use variable(s)     Widget: subgroup_buttons \n")
        
        # Variables that need to be initialized with specific values/number of items
        st.session_state["column_data"] = ["0", "1", "2", "3", "4"]
        st.session_state["KM_data_all"] = pd.DataFrame(columns=["0", "1", "2"])

        # Show a slider for 1 to 5 variables (to prevent crazy number of curves)
        with subgroup_buttons_output:
            variable_number_slider = st.slider(label="Number of variables:", 
                                                min_value=1, max_value=5, step=1, value=1)
        variable_number_slider_handler(variable_number_slider)   
    else:
        # If the user doesn't want to make subgroups, don't show the slider 
        logger.info(f"The user selected: No subgroups     Widget: subgroup_buttons \n")
        subgroup_buttons_output.empty()
        
        # Clear memomory if subgroup were previously made and no longer needed
        if "subgroup_info" in st.session_state:
            del st.session_state["subgroup_info"]
        if "KM_data_all" in st.session_state:
            del st.session_state["KM_data_all"]

###################################################################################################

# Function to display the output of variable_number_slider (widget+output areas)
def variable_number_slider_handler(change):
    
    # Get required variables from the session state
    df_clinical = st.session_state.get("df_clinical")
    df_RNA = st.session_state.get("df_RNA")
    logger = st.session_state.get("logger")
    
    # The only way to make all the widgets independent, their values accesible, and persistent across
    # streamlit re-runs was to repeat the following code (instead of using for loops to dinamically 
    # generate the needed number of repeats)

    #################### Repeat 1
    if change >= 1:
        # Create a list to store the column containers
        st.markdown('<hr style="margin-top: +10px; margin-bottom: +10px; border-width: 5px;">', unsafe_allow_html=True)
        col_1_row_4, col_2_row_4, col_3_row_4 = st.columns([1,1,1])
        col_1_row_5, col_2_row_5 = st.columns([3,1])

        # Show the first widget - dataset dropdown
        with col_1_row_4:
            dataset_dropdown_1 = st.selectbox(label="Select a dataset:", 
                        options=["Click here to select...", "clinical"] + (["RNA"] if df_RNA is not None else []),
                        index=0, key="dataset_dropdown_1")
        
        # When something is selected in the first widget, show two more in the same row
        if dataset_dropdown_1 == "clinical":
            # A dropdown widget if the dataset selected is the clinical one
            with col_2_row_4:
                variable_dropdown_1 = st.selectbox(label="Select a variable:",
                                    options=["Click here to select..."] + list(df_clinical.columns[1:]),
                                    index=0, key="variable_dropdown_1")
            # And a slider to select the number of subgroups to make
            with col_3_row_4:
                subgroup_slider_1 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=5, value=1, step=1, key="subgroup_slider_1") 
        elif dataset_dropdown_1 == "RNA":
            # A searchbox widget if the dataset selected is the RNA one
            with col_2_row_4:
                variable_dropdown_1 = st_searchbox(search_function=search_genes, default="Click here to select...", 
                                label="Type a gene name here", clear_on_submit=False, key="variable_dropdown_1rna")
            # The same slider as above, but putting it inside the if-else prevents it from showing up immediately 
            with col_3_row_4:
                subgroup_slider_1 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=5, value=1, step=1, key="subgroup_slider_1") 
        
        # Encapsulate these checks for the two additional widgets as they don't exist until a dataset is selected
        try:
            # When a variable is selected in either dropdown or searchbox, make a bar chart or histogram
            if variable_dropdown_1 != "Click here to select...":
                # There is a function that creates the appropriate plot, we just need the variable and repeat number
                variable_figure_1 = variables_selection_handler(variable_dropdown_1, 1)
                
                # Since we need to apply immediately the 0/1 event tags, handle when they have not been specified
                with col_1_row_5:
                    if variable_figure_1 is None:
                        st.warning("First select the values to label as 0 and 1 (No event, event)!!")
                        st.stop()
                    else:
                        st.altair_chart(variable_figure_1)
                
                # Prepare the label options to make subgroups by either tags or float ranges
                if dataset_dropdown_1 == "clinical":
                    if df_clinical[variable_dropdown_1].dtype == "object": 
                        # Scenario 1 - a real text column
                        subgrouping_options_1 = ["tags"] + list(df_clinical[variable_dropdown_1].unique())
                    else:
                        # Display correct widgets according to the plot made (see 8 scenarios considered)
                        # We need to check the number of unique values and type of data in the column
                        nonobject_data_check_1 = df_clinical[variable_dropdown_1].dropna().copy()
                        n_unique_1 = nonobject_data_check_1.nunique()
                        is_float_dtype_1 = pd.api.types.is_float_dtype(nonobject_data_check_1)

                        # Scenarios 2, 3, 6, 7: >10 unique values, int or float with or without NaN
                        if n_unique_1 > 10:
                            # If all values are already integers, no conversion is needed (scenario 2)
                            if all(isinstance(value, int) for value in nonobject_data_check_1):
                                step_1 = 1
                            # Check for integers that were transformed to floats due to NaNs (scenario 6)
                            elif is_float_dtype_1 and all(value.is_integer() for value in nonobject_data_check_1):
                                # Convert the elements back to integers and the column dtype to integer too
                                nonobject_data_check_1 = nonobject_data_check_1.apply(lambda x: int(x) if np.isfinite(x) else x)
                                nonobject_data_check_1 = nonobject_data_check_1.astype('Int64')
                                step_1 = 1
                            # If all values were truly floats, no conversion is needed (scenario 3 and 7)
                            else:
                                step_1 = 0.1
                            
                            # Generate the info for the range widget for both int or float
                            subgrouping_options_1 = ["ranges"] + [step_1] + nonobject_data_check_1.agg(["min", "max"]).tolist()
                        # Scenarios 4, 5, 8, 9: <10 unique values, int or float with or without NaN
                        elif n_unique_1 < 10:
                            # Transform column dtype to object and values to strings
                            nonobject_data_check_1 = nonobject_data_check_1.astype('object')
                            nonobject_data_check_1 = nonobject_data_check_1.apply(lambda x: str(x))
                            
                            # Check if all non-NaN values end with '.0' (they were likely integers-Scenario 8)
                            if all(str(x).endswith('.0') for x in nonobject_data_check_1):
                                # Remove '.0' from these values
                                nonobject_data_check_1 = nonobject_data_check_1.apply(lambda x: str(x)[:-2] if str(x).endswith('.0') else str(x))
                            
                            # Generate the info for the multiselect widget for both int or float
                            subgrouping_options_1 = ["tags"] + list(nonobject_data_check_1.unique())
                # RNA data is always float numbers
                elif dataset_dropdown_1 == "RNA":
                    subgrouping_options_1 = ["ranges"] + [0.1] + df_RNA[variable_dropdown_1].agg(["min", "max"]).tolist()
            
            # When the slider is changed, show additional widgets to specify the subgroups
            if subgroup_slider_1 == 1:
                with col_2_row_5:
                    st.warning("For 1 group select -None-")
            else:
                # Show widgets to make 2 to 5 subgroups (either based on tags or number ranges)
                if subgroup_slider_1 >= 2:
                    if subgrouping_options_1[0] == "tags":
                        with col_2_row_5:
                            subgroup_1_1 = st.multiselect(label="Subgroup 1:", options=subgrouping_options_1[1:])
                            subgroup_1_2 = st.multiselect(label="Subgroup 2:", options=subgrouping_options_1[1:])
                    else:
                        with col_2_row_5:
                            subgroup_1_1 = st.slider(label="Subgroup 1:", step=subgrouping_options_1[1],
                                                    min_value=subgrouping_options_1[2],
                                                    max_value=subgrouping_options_1[3],
                                                    value=(subgrouping_options_1[2], subgrouping_options_1[3]))
                            subgroup_1_2 = st.slider(label="Subgroup 2:", step=subgrouping_options_1[1],
                                                    min_value=subgrouping_options_1[2],
                                                    max_value=subgrouping_options_1[3],
                                                    value=(subgrouping_options_1[2], subgrouping_options_1[3]))
                if subgroup_slider_1 >= 3:
                    if subgrouping_options_1[0] == "tags":
                        with col_2_row_5:
                            subgroup_1_3 = st.multiselect(label="Subgroup 3:", options=subgrouping_options_1[1:])
                    else:
                        with col_2_row_5:
                            subgroup_1_3 = st.slider(label="Subgroup 3:", step=subgrouping_options_1[1],
                                                    min_value=subgrouping_options_1[2],
                                                    max_value=subgrouping_options_1[3],
                                                    value=(subgrouping_options_1[2], subgrouping_options_1[3]))
                if subgroup_slider_1 >= 4:
                    if subgrouping_options_1[0] == "tags":
                        with col_2_row_5:
                            subgroup_1_4 = st.multiselect(label="Subgroup 4:", options=subgrouping_options_1[1:])
                    else:
                        with col_2_row_5:
                            subgroup_1_4 = st.slider(label="Subgroup 4:", step=subgrouping_options_1[1],
                                                    min_value=subgrouping_options_1[2],
                                                    max_value=subgrouping_options_1[3],
                                                    value=(subgrouping_options_1[2], subgrouping_options_1[3]))
                if subgroup_slider_1 >= 5:
                    if subgrouping_options_1[0] == "tags":
                        with col_2_row_5:
                            subgroup_1_5 = st.multiselect(label="Subgroup 5:", options=subgrouping_options_1[1:])
                    else:
                        with col_2_row_5:
                            subgroup_1_5 = st.slider(label="Subgroup 5:", step=subgrouping_options_1[1],
                                                    min_value=subgrouping_options_1[2],
                                                    max_value=subgrouping_options_1[3],
                                                    value=(subgrouping_options_1[2], subgrouping_options_1[3]))
        except NameError:
            pass
    ###############
        
    #################### Repeat 2
    if change >= 2:
        # Create a list to store the column containers
        st.markdown('<hr style="margin-top: +10px; margin-bottom: +10px; border-width: 5px;">', unsafe_allow_html=True)
        col_1_row_6, col_2_row_6, col_3_row_6 = st.columns([1,1,1])
        col_1_row_7, col_2_row_7 = st.columns([3,1])

        # Show the first widget - dataset dropdown
        with col_1_row_6:
            dataset_dropdown_2 = st.selectbox(label="Select a dataset:", 
                        options=["Click here to select...", "clinical"] + (["RNA"] if df_RNA is not None else []),
                        index=0, key="dataset_dropdown_2")
        
        # When something is selected in the first widget, show two more in the same row
        if dataset_dropdown_2 == "clinical":
            # A dropdown widget if the dataset selected is the clinical one
            with col_2_row_6:
                variable_dropdown_2 = st.selectbox(label="Select a variable:",
                                    options=["Click here to select..."] + list(df_clinical.columns[1:]),
                                    index=0, key="variable_dropdown_2")
            # And a slider to select the number of subgroups to make
            with col_3_row_6:
                subgroup_slider_2 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=5, value=1, step=1, key="subgroup_slider_2") 
        elif dataset_dropdown_2 == "RNA":
            # A searchbox widget if the dataset selected is the RNA one
            with col_2_row_6:
                variable_dropdown_2 = st_searchbox(search_function=search_genes, default="Click here to select...", 
                                label="Type a gene name here", clear_on_submit=False, key="variable_dropdown_2rna")
            # The same slider as above, but putting it inside the if-else prevents it from showing up immediately 
            with col_3_row_6:
                subgroup_slider_2 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=5, value=1, step=1, key="subgroup_slider_2") 
        
        # Encapsulate these checks for the two additional widgets as they don't exist until a dataset is selected
        try:
            # When a variable is selected in either dropdown or searchbox, make a bar chart or histogram
            if variable_dropdown_2 != "Click here to select...":
                # There is a function that creates the appropriate plot, we just need the variable and repeat number
                variable_figure_2 = variables_selection_handler(variable_dropdown_2, 2)
                
                # Since we need to apply immediately the 0/1 event tags, handle when they have not been specified
                with col_1_row_7:
                    if variable_figure_2 is None:
                        st.warning("First select the values to label as 0 and 1 (No event, event)!!")
                        st.stop()
                    else:
                        st.altair_chart(variable_figure_2)
                
                # Prepare the label options to make subgroups by either tags or float ranges
                if dataset_dropdown_2 == "clinical":
                    if df_clinical[variable_dropdown_2].dtype == "object": 
                        # Scenario 1 - a real text column
                        subgrouping_options_2 = ["tags"] + list(df_clinical[variable_dropdown_2].unique())
                    else:
                        # Display correct widgets according to the plot made (see 8 scenarios considered)
                        # We need to check the number of unique values and type of data in the column
                        nonobject_data_check_2 = df_clinical[variable_dropdown_2].dropna().copy()
                        n_unique_2 = nonobject_data_check_2.nunique()
                        is_float_dtype_2 = pd.api.types.is_float_dtype(nonobject_data_check_2)

                        # Scenarios 2, 3, 6, 7: >10 unique values, int or float with or without NaN
                        if n_unique_2 > 10:
                            # If all values are already integers, no conversion is needed (scenario 2)
                            if all(isinstance(value, int) for value in nonobject_data_check_2):
                                step_2 = 1
                            # Check for integers that were transformed to floats due to NaNs (scenario 6)
                            elif is_float_dtype_2 and all(value.is_integer() for value in nonobject_data_check_2):
                                # Convert the elements back to integers and the column dtype to integer too
                                nonobject_data_check_2 = nonobject_data_check_2.apply(lambda x: int(x) if np.isfinite(x) else x)
                                nonobject_data_check_2 = nonobject_data_check_2.astype('Int64')
                                step_2 = 1
                            # If all values were truly floats, no conversion is needed (scenario 3 and 7)
                            else:
                                step_2 = 0.1
                            
                            # Generate the info for the range widget for both int or float
                            subgrouping_options_2 = ["ranges"] + [step_2] + nonobject_data_check_2.agg(["min", "max"]).tolist()
                        # Scenarios 4, 5, 8, 9: <10 unique values, int or float with or without NaN
                        elif n_unique_2 < 10:
                            # Transform column dtype to object and values to strings
                            nonobject_data_check_2 = nonobject_data_check_2.astype('object')
                            nonobject_data_check_2 = nonobject_data_check_2.apply(lambda x: str(x))
                            
                            # Check if all non-NaN values end with '.0' (they were likely integers-Scenario 8)
                            if all(str(x).endswith('.0') for x in nonobject_data_check_2):
                                # Remove '.0' from these values
                                nonobject_data_check_2 = nonobject_data_check_2.apply(lambda x: str(x)[:-2] if str(x).endswith('.0') else str(x))
                            
                            # Generate the info for the multiselect widget for both int or float
                            subgrouping_options_2 = ["tags"] + list(nonobject_data_check_2.unique())
                # RNA data is always float numbers
                elif dataset_dropdown_2 == "RNA":
                    subgrouping_options_2 = ["ranges"] + [0.1] + df_RNA[variable_dropdown_2].agg(["min", "max"]).tolist()
            
            # When the slider is changed, show additional widgets to specify the subgroups
            if subgroup_slider_2 == 1:
                with col_2_row_7:
                    st.warning("For 1 group select -None-")
            else:
                # Show widgets to make 2 to 5 subgroups (either based on tags or number ranges)
                if subgroup_slider_2 >= 2:
                    if subgrouping_options_2[0] == "tags":
                        with col_2_row_7:
                            subgroup_2_1 = st.multiselect(label="Subgroup 1:", options=subgrouping_options_2[1:])
                            subgroup_2_2 = st.multiselect(label="Subgroup 2:", options=subgrouping_options_2[1:])
                    else:
                        with col_2_row_7:
                            subgroup_2_1 = st.slider(label="Subgroup 1:", step=subgrouping_options_2[1],
                                                    min_value=subgrouping_options_2[2],
                                                    max_value=subgrouping_options_2[3],
                                                    value=(subgrouping_options_2[2], subgrouping_options_2[3]))
                            subgroup_2_2 = st.slider(label="Subgroup 2:", step=subgrouping_options_2[1],
                                                    min_value=subgrouping_options_2[2],
                                                    max_value=subgrouping_options_2[3],
                                                    value=(subgrouping_options_2[2], subgrouping_options_2[3]))
                if subgroup_slider_2 >= 3:
                    if subgrouping_options_2[0] == "tags":
                        with col_2_row_7:
                            subgroup_2_3 = st.multiselect(label="Subgroup 3:", options=subgrouping_options_2[1:])
                    else:
                        with col_2_row_7:
                            subgroup_2_3 = st.slider(label="Subgroup 3:", step=subgrouping_options_2[1],
                                                    min_value=subgrouping_options_2[2],
                                                    max_value=subgrouping_options_2[3],
                                                    value=(subgrouping_options_2[2], subgrouping_options_2[3]))
                if subgroup_slider_2 >= 4:
                    if subgrouping_options_2[0] == "tags":
                        with col_2_row_7:
                            subgroup_2_4 = st.multiselect(label="Subgroup 4:", options=subgrouping_options_2[1:])
                    else:
                        with col_2_row_7:
                            subgroup_2_4 = st.slider(label="Subgroup 4:", step=subgrouping_options_2[1],
                                                    min_value=subgrouping_options_2[2],
                                                    max_value=subgrouping_options_2[3],
                                                    value=(subgrouping_options_2[2], subgrouping_options_2[3]))
                if subgroup_slider_2 >= 5:
                    if subgrouping_options_2[0] == "tags":
                        with col_2_row_7:
                            subgroup_2_5 = st.multiselect(label="Subgroup 5:", options=subgrouping_options_2[1:])
                    else:
                        with col_2_row_7:
                            subgroup_2_5 = st.slider(label="Subgroup 5:", step=subgrouping_options_2[1],
                                                    min_value=subgrouping_options_2[2],
                                                    max_value=subgrouping_options_2[3],
                                                    value=(subgrouping_options_2[2], subgrouping_options_2[3]))
        except NameError:
            pass
    ###############

    #################### Repeat 3
    if change >= 3:
        # Create a list to store the column containers
        st.markdown('<hr style="margin-top: +10px; margin-bottom: +10px; border-width: 5px;">', unsafe_allow_html=True)
        col_1_row_8, col_2_row_8, col_3_row_8 = st.columns([1,1,1])
        col_1_row_9, col_2_row_9 = st.columns([3,1])

        # Show the first widget - dataset dropdown
        with col_1_row_8:
            dataset_dropdown_3 = st.selectbox(label="Select a dataset:", 
                        options=["Click here to select...", "clinical"] + (["RNA"] if df_RNA is not None else []),
                        index=0, key="dataset_dropdown_3")
        
        # When something is selected in the first widget, show two more in the same row
        if dataset_dropdown_3 == "clinical":
            # A dropdown widget if the dataset selected is the clinical one
            with col_2_row_8:
                variable_dropdown_3 = st.selectbox(label="Select a variable:",
                                    options=["Click here to select..."] + list(df_clinical.columns[1:]),
                                    index=0, key="variable_dropdown_3")
            # And a slider to select the number of subgroups to make
            with col_3_row_8:
                subgroup_slider_3 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=5, value=1, step=1, key="subgroup_slider_3") 
        elif dataset_dropdown_3 == "RNA":
            # A searchbox widget if the dataset selected is the RNA one
            with col_2_row_8:
                variable_dropdown_3 = st_searchbox(search_function=search_genes, default="Click here to select...", 
                                label="Type a gene name here", clear_on_submit=False, key="variable_dropdown_3rna")
            # The same slider as above, but putting it inside the if-else prevents it from showing up immediately 
            with col_3_row_8:
                subgroup_slider_3 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=5, value=1, step=1, key="subgroup_slider_3") 
        
        # Encapsulate these checks for the two additional widgets as they don't exist until a dataset is selected
        try:
            # When a variable is selected in either dropdown or searchbox, make a bar chart or histogram
            if variable_dropdown_3 != "Click here to select...":
                # There is a function that creates the appropriate plot, we just need the variable and repeat number
                variable_figure_3 = variables_selection_handler(variable_dropdown_3, 3)
                
                # Since we need to apply immediately the 0/1 event tags, handle when they have not been specified
                with col_1_row_9:
                    if variable_figure_3 is None:
                        st.warning("First select the values to label as 0 and 1 (No event, event)!!")
                        st.stop()
                    else:
                        st.altair_chart(variable_figure_3)
                
                # Prepare the label options to make subgroups by either tags or float ranges
                if dataset_dropdown_3 == "clinical":
                    if df_clinical[variable_dropdown_3].dtype == "object": 
                        subgrouping_options_3 = ["tags"] + list(df_clinical[variable_dropdown_3].unique())
                    else:
                        nonobject_data_check_3 = df_clinical[variable_dropdown_3].dropna().copy()
                        n_unique_3 = nonobject_data_check_3.nunique()
                        is_float_dtype_3 = pd.api.types.is_float_dtype(nonobject_data_check_3)

                        if n_unique_3 > 10:
                            if all(isinstance(value, int) for value in nonobject_data_check_3):
                                step_3 = 1
                            elif is_float_dtype_3 and all(value.is_integer() for value in nonobject_data_check_3):
                                nonobject_data_check_3 = nonobject_data_check_3.apply(lambda x: int(x) if np.isfinite(x) else x)
                                nonobject_data_check_3 = nonobject_data_check_3.astype('Int64')
                                step_3 = 1
                            else:
                                step_3 = 0.1
                            
                            subgrouping_options_3 = ["ranges"] + [step_3] + nonobject_data_check_3.agg(["min", "max"]).tolist()
                        elif n_unique_3 < 10:
                            nonobject_data_check_3 = nonobject_data_check_3.astype('object')
                            nonobject_data_check_3 = nonobject_data_check_3.apply(lambda x: str(x))
                            
                            if all(str(x).endswith('.0') for x in nonobject_data_check_3):
                                nonobject_data_check_3 = nonobject_data_check_3.apply(lambda x: str(x)[:-2] if str(x).endswith('.0') else str(x))
                            
                            subgrouping_options_3 = ["tags"] + list(nonobject_data_check_3.unique())
                elif dataset_dropdown_3 == "RNA":
                    subgrouping_options_3 = ["ranges"] + [0.1] + df_RNA[variable_dropdown_3].agg(["min", "max"]).tolist()
            
            if subgroup_slider_3 == 1:
                with col_2_row_9:
                    st.warning("For 1 group select -None-")
            else:
                if subgroup_slider_3 >= 2:
                    if subgrouping_options_3[0] == "tags":
                        with col_2_row_9:
                            subgroup_3_1 = st.multiselect(label="Subgroup 1:", options=subgrouping_options_3[1:])
                            subgroup_3_2 = st.multiselect(label="Subgroup 2:", options=subgrouping_options_3[1:])
                    else:
                        with col_2_row_9:
                            subgroup_3_1 = st.slider(label="Subgroup 1:", step=subgrouping_options_3[1],
                                                    min_value=subgrouping_options_3[2],
                                                    max_value=subgrouping_options_3[3],
                                                    value=(subgrouping_options_3[2], subgrouping_options_3[3]))
                            subgroup_3_2 = st.slider(label="Subgroup 2:", step=subgrouping_options_3[1],
                                                    min_value=subgrouping_options_3[2],
                                                    max_value=subgrouping_options_3[3],
                                                    value=(subgrouping_options_3[2], subgrouping_options_3[3]))
            if subgroup_slider_3 >= 3:
                if subgrouping_options_3[0] == "tags":
                    with col_2_row_9:
                        subgroup_3_3 = st.multiselect(label="Subgroup 3:", options=subgrouping_options_3[1:])
                else:
                    with col_2_row_9:
                        subgroup_3_3 = st.slider(label="Subgroup 3:", step=subgrouping_options_3[1],
                                                min_value=subgrouping_options_3[2],
                                                max_value=subgrouping_options_3[3],
                                                value=(subgrouping_options_3[2], subgrouping_options_3[3]))
                if subgroup_slider_3 >= 4:
                    if subgrouping_options_3[0] == "tags":
                        with col_2_row_9:
                            subgroup_3_4 = st.multiselect(label="Subgroup 4:", options=subgrouping_options_3[1:])
                    else:
                        with col_2_row_9:
                            subgroup_3_4 = st.slider(label="Subgroup 4:", step=subgrouping_options_3[1],
                                                    min_value=subgrouping_options_3[2],
                                                    max_value=subgrouping_options_3[3],
                                                    value=(subgrouping_options_3[2], subgrouping_options_3[3]))
                if subgroup_slider_3 >= 5:
                    if subgrouping_options_3[0] == "tags":
                        with col_2_row_9:
                            subgroup_3_5 = st.multiselect(label="Subgroup 5:", options=subgrouping_options_3[1:])
                    else:
                        with col_2_row_9:
                            subgroup_3_5 = st.slider(label="Subgroup 5:", step=subgrouping_options_3[1],
                                                    min_value=subgrouping_options_3[2],
                                                    max_value=subgrouping_options_3[3],
                                                    value=(subgrouping_options_3[2], subgrouping_options_3[3]))
        except NameError:
            pass
    ###############

    #################### Repeat 4
    if change >= 4:
        # Create a list to store the column containers
        st.markdown('<hr style="margin-top: +10px; margin-bottom: +10px; border-width: 5px;">', unsafe_allow_html=True)
        col_1_row_10, col_2_row_10, col_3_row_10 = st.columns([1,1,1])
        col_1_row_11, col_2_row_11 = st.columns([3,1])

        # Show the first widget - dataset dropdown
        with col_1_row_10:
            dataset_dropdown_4 = st.selectbox(label="Select a dataset:", 
                        options=["Click here to select...", "clinical"] + (["RNA"] if df_RNA is not None else []),
                        index=0, key="dataset_dropdown_4")
        
        # When something is selected in the first widget, show two more in the same row
        if dataset_dropdown_4 == "clinical":
            # A dropdown widget if the dataset selected is the clinical one
            with col_2_row_10:
                variable_dropdown_4 = st.selectbox(label="Select a variable:",
                                    options=["Click here to select..."] + list(df_clinical.columns[1:]),
                                    index=0, key="variable_dropdown_4")
            # And a slider to select the number of subgroups to make
            with col_3_row_10:
                subgroup_slider_4 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=5, value=1, step=1, key="subgroup_slider_4") 
        elif dataset_dropdown_4 == "RNA":
            # A searchbox widget if the dataset selected is the RNA one
            with col_2_row_10:
                variable_dropdown_4 = st_searchbox(search_function=search_genes, default="Click here to select...", 
                                label="Type a gene name here", clear_on_submit=False, key="variable_dropdown_4rna")
            # The same slider as above, but putting it inside the if-else prevents it from showing up immediately 
            with col_3_row_10:
                subgroup_slider_4 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=5, value=1, step=1, key="subgroup_slider_4") 
        
        # Encapsulate these checks for the two additional widgets as they don't exist until a dataset is selected
        try:
            # When a variable is selected in either dropdown or searchbox, make a bar chart or histogram
            if variable_dropdown_4 != "Click here to select...":
                # There is a function that creates the appropriate plot, we just need the variable and repeat number
                variable_figure_4 = variables_selection_handler(variable_dropdown_4, 4)
                
                # Since we need to apply immediately the 0/1 event tags, handle when they have not been specified
                with col_1_row_11:
                    if variable_figure_4 is None:
                        st.warning("First select the values to label as 0 and 1 (No event, event)!!")
                        st.stop()
                    else:
                        st.altair_chart(variable_figure_4)
                
                # Prepare the label options to make subgroups by either tags or float ranges
                if dataset_dropdown_4 == "clinical":
                    if df_clinical[variable_dropdown_4].dtype == "object": 
                        subgrouping_options_4 = ["tags"] + list(df_clinical[variable_dropdown_4].unique())
                    else:
                        nonobject_data_check_4 = df_clinical[variable_dropdown_4].dropna().copy()
                        n_unique_4 = nonobject_data_check_4.nunique()
                        is_float_dtype_4 = pd.api.types.is_float_dtype(nonobject_data_check_4)

                        if n_unique_4 > 10:
                            if all(isinstance(value, int) for value in nonobject_data_check_4):
                                step_4 = 1
                            elif is_float_dtype_4 and all(value.is_integer() for value in nonobject_data_check_4):
                                nonobject_data_check_4 = nonobject_data_check_4.apply(lambda x: int(x) if np.isfinite(x) else x)
                                nonobject_data_check_4 = nonobject_data_check_4.astype('Int64')
                                step_4 = 1
                            else:
                                step_4 = 0.1
                            
                            subgrouping_options_4 = ["ranges"] + [step_4] + nonobject_data_check_4.agg(["min", "max"]).tolist()
                        elif n_unique_4 < 10:
                            nonobject_data_check_4 = nonobject_data_check_4.astype('object')
                            nonobject_data_check_4 = nonobject_data_check_4.apply(lambda x: str(x))
                            
                            if all(str(x).endswith('.0') for x in nonobject_data_check_4):
                                nonobject_data_check_4 = nonobject_data_check_4.apply(lambda x: str(x)[:-2] if str(x).endswith('.0') else str(x))
                            
                            subgrouping_options_4 = ["tags"] + list(nonobject_data_check_4.unique())
                elif dataset_dropdown_4 == "RNA":
                    subgrouping_options_4 = ["ranges"] + [0.1] + df_RNA[variable_dropdown_4].agg(["min", "max"]).tolist()
            
            if subgroup_slider_4 == 1:
                with col_2_row_11:
                    st.warning("For 1 group select -None-")
            else:
                if subgroup_slider_4 >= 2:
                    if subgrouping_options_4[0] == "tags":
                        with col_2_row_11:
                            subgroup_4_1 = st.multiselect(label="Subgroup 1:", options=subgrouping_options_4[1:])
                            subgroup_4_2 = st.multiselect(label="Subgroup 2:", options=subgrouping_options_4[1:])
                    else:
                        with col_2_row_11:
                            subgroup_4_1 = st.slider(label="Subgroup 1:", step=subgrouping_options_4[1],
                                                    min_value=subgrouping_options_4[2],
                                                    max_value=subgrouping_options_4[3],
                                                    value=(subgrouping_options_4[2], subgrouping_options_4[3]))
                            subgroup_4_2 = st.slider(label="Subgroup 2:", step=subgrouping_options_4[1],
                                                    min_value=subgrouping_options_4[2],
                                                    max_value=subgrouping_options_4[3],
                                                    value=(subgrouping_options_4[2], subgrouping_options_4[3]))
            if subgroup_slider_4 >= 3:
                if subgrouping_options_4[0] == "tags":
                    with col_2_row_11:
                        subgroup_4_3 = st.multiselect(label="Subgroup 3:", options=subgrouping_options_4[1:])
                else:
                    with col_2_row_11:
                        subgroup_4_3 = st.slider(label="Subgroup 3:", step=subgrouping_options_4[1],
                                                min_value=subgrouping_options_4[2],
                                                max_value=subgrouping_options_4[3],
                                                value=(subgrouping_options_4[2], subgrouping_options_4[3]))
                if subgroup_slider_4 >= 4:
                    if subgrouping_options_4[0] == "tags":
                        with col_2_row_11:
                            subgroup_4_4 = st.multiselect(label="Subgroup 4:", options=subgrouping_options_4[1:])
                    else:
                        with col_2_row_11:
                            subgroup_4_4 = st.slider(label="Subgroup 4:", step=subgrouping_options_4[1],
                                                    min_value=subgrouping_options_4[2],
                                                    max_value=subgrouping_options_4[3],
                                                    value=(subgrouping_options_4[2], subgrouping_options_4[3]))
                if subgroup_slider_4 >= 5:
                    if subgrouping_options_4[0] == "tags":
                        with col_2_row_11:
                            subgroup_4_5 = st.multiselect(label="Subgroup 5:", options=subgrouping_options_4[1:])
                    else:
                        with col_2_row_11:
                            subgroup_4_5 = st.slider(label="Subgroup 5:", step=subgrouping_options_4[1],
                                                    min_value=subgrouping_options_4[2],
                                                    max_value=subgrouping_options_4[3],
                                                    value=(subgrouping_options_4[2], subgrouping_options_4[3]))
        except NameError:
            pass
    ###############

    #################### Repeat 5
    if change >= 5:
        # Create a list to store the column containers
        st.markdown('<hr style="margin-top: +10px; margin-bottom: +10px; border-width: 5px;">', unsafe_allow_html=True)
        col_1_row_12, col_2_row_12, col_3_row_12 = st.columns([1,1,1])
        col_1_row_13, col_2_row_13 = st.columns([3,1])

        # Show the first widget - dataset dropdown
        with col_1_row_12:
            dataset_dropdown_5 = st.selectbox(label="Select a dataset:", 
                        options=["Click here to select...", "clinical"] + (["RNA"] if df_RNA is not None else []),
                        index=0, key="dataset_dropdown_5")
        
        # When something is selected in the first widget, show two more in the same row
        if dataset_dropdown_5 == "clinical":
            # A dropdown widget if the dataset selected is the clinical one
            with col_2_row_12:
                variable_dropdown_5 = st.selectbox(label="Select a variable:",
                                    options=["Click here to select..."] + list(df_clinical.columns[1:]),
                                    index=0, key="variable_dropdown_5")
            # And a slider to select the number of subgroups to make
            with col_3_row_12:
                subgroup_slider_5 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=5, value=1, step=1, key="subgroup_slider_5") 
        elif dataset_dropdown_5 == "RNA":
            # A searchbox widget if the dataset selected is the RNA one
            with col_2_row_12:
                variable_dropdown_5 = st_searchbox(search_function=search_genes, default="Click here to select...", 
                                label="Type a gene name here", clear_on_submit=False, key="variable_dropdown_5rna")
            # The same slider as above, but putting it inside the if-else prevents it from showing up immediately 
            with col_3_row_12:
                subgroup_slider_5 = st.slider(label="Number of subgroups:",
                                min_value=1, max_value=5, value=1, step=1, key="subgroup_slider_5") 
        
        # Encapsulate these checks for the two additional widgets as they don't exist until a dataset is selected
        try:
            # When a variable is selected in either dropdown or searchbox, make a bar chart or histogram
            if variable_dropdown_5 != "Click here to select...":
                # There is a function that creates the appropriate plot, we just need the variable and repeat number
                variable_figure_5 = variables_selection_handler(variable_dropdown_5, 5)
                
                # Since we need to apply immediately the 0/1 event tags, handle when they have not been specified
                with col_1_row_13:
                    if variable_figure_5 is None:
                        st.warning("First select the values to label as 0 and 1 (No event, event)!!")
                        st.stop()
                    else:
                        st.altair_chart(variable_figure_5)
                
                # Prepare the label options to make subgroups by either tags or float ranges
                if dataset_dropdown_5 == "clinical":
                    if df_clinical[variable_dropdown_5].dtype == "object": 
                        subgrouping_options_5 = ["tags"] + list(df_clinical[variable_dropdown_5].unique())
                    else:
                        nonobject_data_check_5 = df_clinical[variable_dropdown_5].dropna().copy()
                        n_unique_5 = nonobject_data_check_5.nunique()
                        is_float_dtype_5 = pd.api.types.is_float_dtype(nonobject_data_check_5)

                        if n_unique_5 > 10:
                            if all(isinstance(value, int) for value in nonobject_data_check_5):
                                step_5 = 1
                            elif is_float_dtype_5 and all(value.is_integer() for value in nonobject_data_check_5):
                                nonobject_data_check_5 = nonobject_data_check_5.apply(lambda x: int(x) if np.isfinite(x) else x)
                                nonobject_data_check_5 = nonobject_data_check_5.astype('Int64')
                                step_5 = 1
                            else:
                                step_5 = 0.1
                            
                            subgrouping_options_5 = ["ranges"] + [step_5] + nonobject_data_check_5.agg(["min", "max"]).tolist()
                        elif n_unique_5 < 10:
                            nonobject_data_check_5 = nonobject_data_check_5.astype('object')
                            nonobject_data_check_5 = nonobject_data_check_5.apply(lambda x: str(x))
                            
                            if all(str(x).endswith('.0') for x in nonobject_data_check_5):
                                nonobject_data_check_5 = nonobject_data_check_5.apply(lambda x: str(x)[:-2] if str(x).endswith('.0') else str(x))
                            
                            subgrouping_options_5 = ["tags"] + list(nonobject_data_check_5.unique())
                elif dataset_dropdown_5 == "RNA":
                    subgrouping_options_5 = ["ranges"] + [0.1] + df_RNA[variable_dropdown_5].agg(["min", "max"]).tolist()
            
            if subgroup_slider_5 == 1:
                with col_2_row_13:
                    st.warning("For 1 group select -None-")
            else:
                if subgroup_slider_5 >= 2:
                    if subgrouping_options_5[0] == "tags":
                        with col_2_row_13:
                            subgroup_5_1 = st.multiselect(label="Subgroup 1:", options=subgrouping_options_5[1:])
                            subgroup_5_2 = st.multiselect(label="Subgroup 2:", options=subgrouping_options_5[1:])
                    else:
                        with col_2_row_13:
                            subgroup_5_1 = st.slider(label="Subgroup 1:", step=subgrouping_options_5[1],
                                                    min_value=subgrouping_options_5[2],
                                                    max_value=subgrouping_options_5[3],
                                                    value=(subgrouping_options_5[2], subgrouping_options_5[3]))
                            subgroup_5_2 = st.slider(label="Subgroup 2:", step=subgrouping_options_5[1],
                                                    min_value=subgrouping_options_5[2],
                                                    max_value=subgrouping_options_5[3],
                                                    value=(subgrouping_options_5[2], subgrouping_options_5[3]))
            if subgroup_slider_5 >= 3:
                if subgrouping_options_5[0] == "tags":
                    with col_2_row_13:
                        subgroup_5_3 = st.multiselect(label="Subgroup 3:", options=subgrouping_options_5[1:])
                else:
                    with col_2_row_13:
                        subgroup_5_3 = st.slider(label="Subgroup 3:", step=subgrouping_options_5[1],
                                                min_value=subgrouping_options_5[2],
                                                max_value=subgrouping_options_5[3],
                                                value=(subgrouping_options_5[2], subgrouping_options_5[3]))
                if subgroup_slider_5 >= 4:
                    if subgrouping_options_5[0] == "tags":
                        with col_2_row_13:
                            subgroup_5_4 = st.multiselect(label="Subgroup 4:", options=subgrouping_options_5[1:])
                    else:
                        with col_2_row_13:
                            subgroup_5_4 = st.slider(label="Subgroup 4:", step=subgrouping_options_5[1],
                                                    min_value=subgrouping_options_5[2],
                                                    max_value=subgrouping_options_5[3],
                                                    value=(subgrouping_options_5[2], subgrouping_options_5[3]))
                if subgroup_slider_5 >= 5:
                    if subgrouping_options_5[0] == "tags":
                        with col_2_row_13:
                            subgroup_5_5 = st.multiselect(label="Subgroup 5:", options=subgrouping_options_5[1:])
                    else:
                        with col_2_row_13:
                            subgroup_5_5 = st.slider(label="Subgroup 5:", step=subgrouping_options_5[1],
                                                    min_value=subgrouping_options_5[2],
                                                    max_value=subgrouping_options_5[3],
                                                    value=(subgrouping_options_5[2], subgrouping_options_5[3]))
        except NameError:
            pass
    ###############

    # Create an empty dictionary for each subgroup
    subgroup_info = [{"df": None, "column": None, 
                    "subgroup_1": None, "subgroup_2": None, "subgroup_3": None,
                    "subgroup_4": None, "subgroup_5": None} for _ in range(5)]

    # Update the subgroup information based on the dropdown values
    for i in range(5):
        try:
            subgroup_info[i]["df"] = locals()[f"dataset_dropdown_{i+1}"]
            subgroup_info[i]["column"] = locals()[f"variable_dropdown_{i+1}"]
            subgroup_info[i]["subgroup_1"] = locals()[f"subgroup_{i+1}_1"]
            subgroup_info[i]["subgroup_2"] = locals()[f"subgroup_{i+1}_2"]
            subgroup_info[i]["subgroup_3"] = locals()[f"subgroup_{i+1}_3"]
            subgroup_info[i]["subgroup_4"] = locals()[f"subgroup_{i+1}_4"]
            subgroup_info[i]["subgroup_5"] = locals()[f"subgroup_{i+1}_5"]
        except:
            continue

    # Save and log the updated subgroup information to the session state
    st.session_state["subgroup_info"] = subgroup_info
    logger.info(f"----------Updated subgroup selections:")
    for i, subgroup in enumerate(subgroup_info):
        logger.info(f"Variable {i+1}, df-{subgroup['df']}, column-{subgroup['column']}")
        logger.info(f"\t Subgroup 1: " + str(subgroup['subgroup_1']) if subgroup['subgroup_1'] is not None else "\t None")
        logger.info(f"\t Subgroup 2: " + str(subgroup['subgroup_2']) if subgroup['subgroup_2'] is not None else "\t None")
        logger.info(f"\t Subgroup 3: " + str(subgroup['subgroup_3']) if subgroup['subgroup_3'] is not None else "\t None")
        logger.info(f"\t Subgroup 4: " + str(subgroup['subgroup_4']) if subgroup['subgroup_4'] is not None else "\t None")
        logger.info(f"\t Subgroup 5: " + str(subgroup['subgroup_5']) if subgroup['subgroup_5'] is not None else "\t None")

###################################################################################################

# Function to search a gene of interest in the RNA dataset (if provided) to make subgroups
def search_genes(searchterm: str) -> List[tuple[str, str]]:
    
    # Assuming st.session_state["gene_list"] is a list of gene names
    suggestions = [gene for gene in st.session_state["gene_list"] if searchterm.lower() in gene.lower()]
    
    # Returning a list of tuples where each tuple contains a label and a value
    return [(gene, gene) for gene in suggestions]

###################################################################################################

# Function to display the output of variables_dropdown and variables_combobox (plots)
# Reminder that this function has to work for both df_clinical and df_RNA
def variables_selection_handler(change, repeat):

    # Get required variables from the session state
    df_clinical = st.session_state.get("df_clinical")
    df_RNA = st.session_state.get("df_RNA")
    time_to_event_selection = st.session_state.get("time_to_event_selection")
    event_observation_selection = st.session_state.get("event_observation_selection")
    event_0 = st.session_state.get("event_0")
    event_1 = st.session_state.get("event_1")
    column_data = st.session_state.get("column_data")
    logger = st.session_state.get("logger")
    alt_colors = st.session_state.get("alt_colors")
    
    # First check that the event labels have been selected
    if event_0 == [] or event_1 == []:
        return 

    ##### Step 01 
    # Apply the 0 and 1 labels to the event observed column and filter those values 
    KM_data_1var = df_clinical.copy()
    for tag in event_0:
        KM_data_1var[event_observation_selection] = KM_data_1var[event_observation_selection].replace(tag, "0")
    for tag in event_1:
        KM_data_1var[event_observation_selection] = KM_data_1var[event_observation_selection].replace(tag, "1")

        # Filter the selected 0/1 event labels, transform column to integers and keep only the common first 3 columns 
        KM_data_1var = KM_data_1var.loc[KM_data_1var[event_observation_selection].isin(["0", "1"])]
        KM_data_1var[event_observation_selection] = KM_data_1var[event_observation_selection].astype(int)

    # Log the current status of KM_data_1var
    logger.info(f"[Subgrouping 1st step] The user selected to label -{str(event_0)}- as 0, and -{str(event_1)}- as 1. \n")
    logger.info(f"[Subgrouping 1st step] Apply 0/1 labels to column {event_observation_selection} on KM_data_1var: \n {KM_data_1var.iloc[:15, :10].to_string()} \n")
    logger.info(f"[Subgrouping 1st step] Data types of KM_data_1var columns: \n {KM_data_1var.dtypes.to_string()} \n\n")
    
    ##### Step 02
    # Look for the selected column in either df, as it is not specified within this function
    if change in df_clinical.columns:
        # Keep only the working columns, log it and extract the column to plot the values
        KM_data_1var = KM_data_1var[["PATIENT_ID", time_to_event_selection, event_observation_selection, change]]          
        logger.info(f"[Subgrouping 2nd step] The column {change} -{KM_data_1var.dtypes[change]} dtype- from df_clinical was selected to make subgroups. \n")
        column_data[repeat - 1] = KM_data_1var[change].copy()
        
    # If the column is in df_RNA, joining is required to combine it with the clinical columns
    elif df_RNA is not None and change in df_RNA.columns:
        # Keep only the working columns from both dfs, log it and extract the column to plot the values
        KM_data_1var = KM_data_1var[["PATIENT_ID", time_to_event_selection, event_observation_selection]]
        df_RNA2 = df_RNA[["PATIENT_ID", change]].copy()
        KM_data_1var = KM_data_1var.merge(df_RNA2, on="PATIENT_ID", how="inner")
        logger.info(f"[Subgrouping 2nd step] The column {change} -{KM_data_1var.dtypes[change]} dtype- from df_RNA was selected to make subgroups. \n")
        column_data[repeat - 1] = KM_data_1var[change].copy()

        # Raise a warning when the patient IDs do not match in the clinical and RNA datasets, causing an empty df
        if len(KM_data_1var) == 0:
            logger.warning(f"[Subgrouping 2nd step] The merge of datasets produced an empty dataframe, it is likely that the IDs may not match. \n")
            st.error("The merge of datasets produced an empty dataframe, check that the patient IDs match in both datasets. \n")
            st.error("You may need to check the data source or pre-process the datasets to make the IDs consistent before using this app.")
            st.stop()

    # Log the current status of KM_data_1var 
    logger.info(f"[Subgrouping 2nd step] Keep relevant columns of KM_data_1var and only rows with 0/1 event labels: \n {KM_data_1var.iloc[:15, :10].to_string()} \n")
    logger.info(f"[Subgrouping 2nd step] Data types of KM_data_1var columns: \n {KM_data_1var.dtypes.to_string()} \n\n")

    ##### Step 03

    # Check the dtype of the column in case some transformation is required
    if column_data[repeat - 1].dtype == "object":
        # Scenario 1 - a real text column, no transformation is required to plot nor in the original KM_data_1var
        alt_data3 = column_data[repeat - 1].value_counts(dropna=False).reset_index()
        alt_data3.columns = [change, "count"]
        alt_data3[change] = alt_data3[change].fillna("NaN")
    else:
        # To treat plot and subset int or float columns, we need to do several checks first (scenario 2-9)
        # If we have more than 10 unique  (int or float), leave it as it is to make a histogram
        # If we have less than 10 unique values (int or float), transform the to object to make a barchart
        # A truly integer column is transformed to float if there are NaNs, so we transform it back if all are .0
        # A truly float column stays as float if there are NaNs
        
        column_data[repeat - 1] = column_data[repeat - 1].dropna()
        n_unique = column_data[repeat - 1].nunique()
        is_float_dtype = pd.api.types.is_float_dtype(column_data[repeat - 1])                        
                            
        # Scenarios 2, 3, 6, 7: >10 unique values, int or float with or without NaN
        if n_unique > 10:
            
            # Check for integers that were transformed to floats due to NaNs (scenario 6)
            if is_float_dtype and all(value.is_integer() for value in column_data[repeat - 1]):
                # Convert the elements back to integers and the column dtype to integer too
                column_data[repeat - 1] = column_data[repeat - 1].apply(lambda x: int(x) if np.isfinite(x) else x)
                column_data[repeat - 1] = column_data[repeat - 1].astype(pd.Int64Dtype().type)
                alt_data3 = pd.DataFrame({change: column_data[repeat - 1]})
            else:
                # If all values are already integers, no conversion is needed (scenario 2)
                # If all values were truly floats, no conversion is needed (scenario 3 and 7)
                alt_data3 = pd.DataFrame({change: column_data[repeat - 1]})
            
            # Update the original column to we keep the transformations when we pass it to the KM analysis
            KM_data_1var[change] = column_data[repeat - 1]
        # Scenarios 4, 5, 8, 9: <10 unique values
        elif n_unique < 10:
            # Transform column dtype to object and all the elements of the column to strings
            column_data[repeat - 1] = column_data[repeat - 1].astype('object')
            column_data[repeat - 1] = column_data[repeat - 1].apply(lambda x: str(x))

            # Check if all non-NaN values end with '.0' (they were likely integers-Scenario 8)
            if all(str(x).endswith('.0') for x in column_data[repeat - 1]):
                # Remove '.0' from these values
                column_data[repeat - 1] = column_data[repeat - 1].apply(lambda x: str(x)[:-2] if str(x).endswith('.0') else str(x))
            
            # Adapt the column data to the format required for the altair plot
            alt_data3 = pd.DataFrame({change: column_data[repeat - 1].unique(), "count": column_data[repeat - 1].value_counts()}).reset_index(drop=True)

            # Update the original column to we keep the transformations when we pass it to the KM analysis
            KM_data_1var[change] = column_data[repeat - 1]

    # Check the type of plot to make based on alt_data3
    if alt_data3[change].dtype == "object":
        # Make a bar chart for text columns showing the counts for unique values
        chart3 = alt.Chart(alt_data3).mark_bar().encode(
            alt.X(change, type="nominal", axis=alt.Axis(labelAngle=-75)),
            alt.Y("count", title="Patients"),
            alt.Color(change, scale=alt.Scale(domain=list(alt_data3[change]), range=alt_colors))
            ).properties(width=500, height=400
            ).configure_axis(labelColor="#3386BD"
            ).configure_legend(disable=True)
    else:
        # Make a histogram of frequencies for numerical columns
        chart3 = alt.Chart(alt_data3).mark_bar(color="#1D8348").encode(
                alt.X(change, type="quantitative", bin=alt.Bin(maxbins=50)),
                alt.Y("count()", title="Patients"),
                ).properties(width=500, height=400
                ).configure_axis(labelColor="#3386BD")
    
    ##### Step 04
    # Pass the information of the current variable KM_data_1var so we can have all variables of interest in KM_data_all
    if repeat == 1:
        # The first variable gets just passed as it is after the steps above
        KM_data_all = KM_data_1var.copy()
    else:
        KM_data_all = st.session_state.get("KM_data_all")
        # We remove the extra time+event columns as we only want to merge the 4th column based on the first
        KM_data_1var = KM_data_1var[["PATIENT_ID", change]]

        # Check if the column is already in KM_data_all
        if change in KM_data_all.columns:
            # Remove the column
            KM_data_all.drop(change, axis=1, inplace=True)
            # Merge the dfs (gets rid of rows with NaN is any of the added columns)
            KM_data_all = KM_data_all.merge(KM_data_1var, how="left", on="PATIENT_ID")
        else:
            # Merge the 4th column of KM_data_1var (gets rid of rows with NaN is any of the added columns)
            KM_data_all = KM_data_all.merge(KM_data_1var, how="left", on="PATIENT_ID")
    
    # Finally, update the state variable
    st.session_state["KM_data_all"] = KM_data_all.copy()

    # Log the KM_data_all as we add/replace variables/columns of interest 
    logger.info(f"[Subgrouping 2nd step] Updated KM_data_all with columns of interest: \n {KM_data_all.iloc[:15, :10].to_string()} \n")

    return chart3

###################################################################################################

# Function to feed the appropriate selections to the KM_analysis function 
def pass_KM_parameters():
    
    # Get the required variables from the session state
    df_clinical = st.session_state.get("df_clinical")
    time_to_event_selection = st.session_state.get("time_to_event_selection")
    event_observation_selection = st.session_state.get("event_observation_selection")
    event_0 = st.session_state.get("event_0")
    event_1 = st.session_state.get("event_1")
    subgroup_buttons_selection = st.session_state.get("subgroup_buttons_selection")
    CI_checkbox = st.session_state.get("CI_checkbox")
    move_labels_checkbox = st.session_state.get("move_labels_checkbox")
    at_risk_checkbox = st.session_state.get("at_risk_checkbox")
    sample_fraction = st.session_state.get("sample_fraction")
    KM_plot_area = st.session_state["widget_and_output_areas"][9]
    logger = st.session_state.get("logger")

    # Handle missing and required widget inputs before trying to do anything
    if time_to_event_selection == "Click here to select...":
        with KM_plot_area:
            KM_plot_area.empty()
            st.warning("First select the time to event column!!")
            st.stop()
    elif event_observation_selection == "Click here to select...":
        with KM_plot_area:
            KM_plot_area.empty()
            st.warning("First select the event observation column!!")
            st.stop()
    elif event_0 == [] or event_1 == []:
        with KM_plot_area:
            KM_plot_area.empty()
            st.warning("First select the values to label as 0 and 1 (No event, event)!!")
            st.stop()
    
    # Check if the time to event column is numerical
    if df_clinical[time_to_event_selection].dtype == "object":
        # First try to convert the column to a numeric type and if it fails we show a warning
        try:
            df_clinical[time_to_event_selection] = pd.to_numeric(df_clinical[time_to_event_selection], errors="coerce")
            df_clinical[time_to_event_selection] = df_clinical[time_to_event_selection].dropna()
        except:
            st.warning("Warning: Time to Event column is not numeric.")
            st.stop()

    # If no subgrouping is required, apply the event tags and pass the data to KM_analysis
    if subgroup_buttons_selection == "None":
        # Apply the selected labels on the event observation column 
        KM_data = df_clinical.copy()
        for tag in event_0:
            KM_data[event_observation_selection] = KM_data[event_observation_selection].replace(tag, "0")
        for tag in event_1:
            KM_data[event_observation_selection] = KM_data[event_observation_selection].replace(tag, "1")            

        # Log the current status of KM_data
        logger.info(f"[No subgroups 1st step] The user selected to label -{str(event_0)}- as 0, and -{str(event_1)}- as 1. \n")
        logger.info(f"[No subgroups 1st step] Apply 0/1 labels to column {event_observation_selection} on KM_data: \n {KM_data.iloc[:15, :10].to_string()} \n")
        logger.info(f"[No subgroups 1st step] Data types of KM_data columns: \n {KM_data.dtypes.to_string()} \n\n")
                
        # Filter out non-desired values and convert column to numbers for the KM Fitter
        KM_data = KM_data[["PATIENT_ID", time_to_event_selection, event_observation_selection]]
        KM_data = KM_data.loc[KM_data[event_observation_selection].isin(["0", "1"])]
        KM_data = KM_data.dropna(subset=[time_to_event_selection])
        KM_data[event_observation_selection] = KM_data[event_observation_selection].astype(int)

        # Log the current status of KM_data
        logger.info(f"[No subgroups 2nd step] Keep relevant columns of KM_data and only rows with 0/1 event labels: \n {KM_data.head(15).to_string()} \n")
        logger.info(f"[No subgroups 2nd step] Data types of KM_data columns: \n {KM_data.dtypes.to_string()} \n\n")
        
        # Pass the input parameters to the KM_analysis function and get back the KM object
        KM_subgroups = []           
        KM_analysis_output = KM_analysis(KM_data, KM_subgroups)
        
        # Make a plot with altair for the KM estimate obtained
        fig, ax = plt.subplots(figsize=(10, 6))
        KM_analysis_output.plot(ci_show=CI_checkbox, legend=False,  
                    iloc=slice(0, int(len(KM_analysis_output.survival_function_) * sample_fraction)))
        ax.set_title("Kaplan-Meier Estimates", fontsize=16)
        ax.set_ylabel("Probability", fontsize=14)
        ax.set_xlabel("Time (Months)", fontsize=14)
        ax.grid(color="#000000", linestyle="--", linewidth=0.5, alpha=0.5)
        ax.set_facecolor("#F0F0F0")
        if at_risk_checkbox:
            add_at_risk_counts(KM_analysis_output, labels=["Label"], ax=ax)

    # If subgroups were selected, apply the corresponding tags or ranges
    else:
        # Get the state variables that only exist if this widget option was selected
        subgroup_info = st.session_state.get("subgroup_info")
        KM_data_all = st.session_state.get("KM_data_all")
        variable_repeats = len(KM_data_all.columns) - 3

        # Log the current status of KM_data_working
        logger.info(f"[Subgrouping 3rd step] Dataset KM_data_all before applying subgrouping labels: \n {KM_data_all.head(15).to_string()} \n")
        logger.info(f"[Subgrouping 3rd step] Data types of KM_data_all before applying subgrouping labels: \n {KM_data_all.dtypes.to_string()} \n\n")
        
        # If the subgrouping changes are made multiple times, apply them to a copy of the original df
        KM_data_working = KM_data_all.copy()

        # Create an empty dictionary to store the mapping for each variable to reassign real group names
        correct_group_labels = [{} for i in range(5)]

        # Iterate just through the number of variables selected to save time
        for repeat in range(variable_repeats):

            # Get the info for the corresponding variable (a dictionary with df, column, and subgroups)
            subgroup_dict = subgroup_info[repeat].copy()
            all_tags_selected = []
            
            # Iterate for each subgroup in the dictionary (the 5 were initialized with None)
            for j in range(1, 6):
                # Get the key-value info
                key = f"subgroup_{j}"
                value = subgroup_dict[key]
                
                # Skip the ones that were not selected/not contain anything
                if value is None:
                    continue
                
                # Check if the subgrouping column is numerical (a tuple of ranges was created)
                if isinstance(value, tuple):
                    # Create a new column to store the group labels 
                    # (the original is numbers and the new one will be text so we cant directly replace the numbers)
                    if "TextSubgroup" not in KM_data_working.columns:
                        KM_data_working.insert(repeat+4, "TextSubgroup", np.nan)
                    
                    # Get the indices of rows within the range selected
                    numeric_values = pd.to_numeric(KM_data_working.iloc[:, repeat+3], errors="coerce")
                    subgroup_rows = (pd.notnull(numeric_values)) & (numeric_values >= value[0]) & (numeric_values < value[1])

                    # Assign the subgroup label to the matching rows
                    KM_data_working.loc[subgroup_rows, "TextSubgroup"] = key
                    
                    # Add the correct label to the dictionary
                    if isinstance(value[0], int) and isinstance(value[1], int):
                        correct_group_labels[repeat][key] = f"{value[0]} to {value[1]}"
                        log_string = f"Subgroup {j}: {value[0]} to {value[1]}"
                    else:
                        correct_group_labels[repeat][key] = f"{value[0]:.2f} to {value[1]:.2f}"
                        log_string = f"Subgroup {j}: {value[0]:.2f} to {value[1]:.2f}"

                    # Log the ranges corresponding to each subgroup
                    logger.info(f"[Subgrouping 3rd step] Subgrouping labels applied to variable {repeat+1}---> {log_string}")

                    # Remove empty rows on the new column only when there are no more subgroups left
                    if j == 5 or subgroup_dict[f"subgroup_{j+1}"] is None:
                        KM_data_working = KM_data_working[pd.notnull(KM_data_working["TextSubgroup"])]
                        # Delete the original column and rename the new one with the original name
                        KM_data_working.drop(subgroup_dict["column"], axis=1, inplace=True)
                        KM_data_working.rename(columns={"TextSubgroup": subgroup_dict["column"]}, inplace=True)
                else:
                    # Generate a mapping of unique values to group labels
                    element_to_label = {element: key for element in value}
                    all_tags_selected.append(key)

                    # Replace the subgroup elements with the new labels
                    KM_data_working[subgroup_dict["column"]] = KM_data_working[subgroup_dict["column"]].replace(element_to_label)

                    # Save and log the labels applied
                    correct_group_labels[repeat][key] = "+".join(value)
                    logger.info(f"[Subgrouping 3rd step] Subgrouping labels applied to variable {repeat+1}---> Subgroup {j}: {value}")
                    
                    # Filter out rows without any of the selected tags only when we have no more subgroups left
                    if j == 5 or subgroup_dict[f"subgroup_{j+1}"] is None:
                        KM_data_working = KM_data_working[KM_data_working[subgroup_dict["column"]].isin(all_tags_selected)]
                    
        # Log the updated df
        logger.info(f"[Subgrouping 3rd step] Dataset KM_data_working after applying subgrouping labels: \n {KM_data_working.head(15).to_string()} \n")
        ########
        # Once all labels have been applied to each column, make the subgroups

        # Get the column indices for the extra columns
        extra_column_indices = range(3, 3 + variable_repeats)
        extra_column_names = ", ".join(KM_data_working.columns[i] for i in extra_column_indices)
        
        # Get the unique values for each extra column
        extra_column_unique_values = [KM_data_working.iloc[:, i].unique() for i in extra_column_indices]
        
        # Create an empty dictionary to store the subsets
        KM_subgroups = {}
        
        # Generate all possible combinations of unique values from the extra columns
        combinations = [[]]
        for values in extra_column_unique_values:
            combinations = [sublist + [value] for sublist in combinations for value in values]
        
        # Iterate through each combination of unique values
        for combination in combinations:
            # Create a subset of KM_data_working for the current combination
            subset = KM_data_working.copy()
            
            # Filter the rows based on the current combination
            for i, index in enumerate(extra_column_indices):
                subset = subset[subset.iloc[:, index] == combination[i]]
            
            # Add the subset to the KM_subgroups dictionary
            KM_subgroups[tuple(combination)] = subset

        # Log the subgroups created
        logger.info(f"[Subgrouping 3rd step] Subgroups made from the dataset:\n")
        for combination, subgroup in KM_subgroups.items():
            logger.info(f"Subgroup label: {combination}")
            logger.info(f"\n{subgroup.head(10)}\n")
        
        ########
                
        # Finally, pass the input parameters to the KM_analysis function and get back the KM object
        KM_analysis_output = KM_analysis(KM_data_working, KM_subgroups)

        # Reassign the real/correct labels (they are as Group X, and we will correct to the actual tag or range)
        # Iterate through each key-value pair in the KM_analysis_output dictionary
        for old_key, KM_object in list(KM_analysis_output.items()):
            # Create a list to store the new key for this combination
            new_key = []
        
            # Iterate through each element of the old key (a tuple of strings)
            for i, label in enumerate(old_key):
                # Retrieve the correct label from the corresponding dictionary in correct_group_labels
                new_key.append(correct_group_labels[i].get(label, label))
        
            # Convert the new key (list) to a single string
            new_key = ", ".join(new_key)
        
            # Replace the current key with the corrected_key in the KM_analysis_output dictionary
            KM_analysis_output[new_key] = KM_analysis_output.pop(old_key)

        # Plot the estimates of all KMF objects 
        with KM_plot_area:
            KM_plot_area.empty()     
            fig, ax = plt.subplots(figsize=(10, 6))
            for label, KM_object in KM_analysis_output.items():
                KM_object.plot(label=label, ci_show=CI_checkbox, 
                               iloc=slice(0, int(len(KM_object.survival_function_) * sample_fraction)))
            ax.set_title("Kaplan-Meier Estimates", fontsize=16)
            ax.set_ylabel("Probability", fontsize=14)
            ax.set_xlabel("Time (Months)", fontsize=14)
            ax.legend(title=extra_column_names)
            ax.grid(color="#000000", linestyle="--", linewidth=0.5, alpha=0.5)
            ax.set_facecolor("#F0F0F0")
            if move_labels_checkbox:
                ax.legend(title=extra_column_names , bbox_to_anchor=(1.05, 1), loc="upper left")
            if at_risk_checkbox:
                add_at_risk_counts(*KM_analysis_output.values(), labels=list(KM_analysis_output.keys()), ax=ax)
    #########################
            
    # Save the KMF objects made by None/Using variable(s) and the figure diplayed to the session state
    st.session_state["KM_analysis_output"] = KM_analysis_output
    figure_bytes = io.BytesIO()
    plt.savefig(figure_bytes, format="jpg", dpi=600, bbox_inches="tight")
    figure_bytes.seek(0)
    st.session_state["logged_figure"] = figure_bytes

    return plt.gcf()

###################################################################################################

def KM_analysis(KM_data, KM_subgroups):
    
    # Unpack the input parameters provided
    current_time_column = KM_data.columns[1]
    current_event_column = KM_data.columns[2]

    # Get the required variables from the session state
    logger = st.session_state.get("logger")
    subgroup_buttons_selection = st.session_state.get("subgroup_buttons_selection")

    # Use the whole dataset when no groups were made
    if subgroup_buttons_selection == "None":

        # Create a single KaplanMeierFitter object
        KMF_object = KaplanMeierFitter()

        # Generate the plot using the specified columns
        KMF_object.fit(durations=KM_data[current_time_column], event_observed=KM_data[current_event_column])

        # Log part of the curve to verify the data was passed correctly
        logger.info(f"[No Subgroups 3rd step] The KM Fitter succesfully calculated the probabilities and made the plot. \n")
        logger.info(f"[No Subgroups 3rd step] Calculated survival function: \n {KMF_object.survival_function_.head(7).to_string()} \n ... \n {KMF_object.survival_function_.tail(7).to_string()} \n\n")

    # Make a fit for every subset provided (based on the number of groups and subgroups made
    else:
        # Sort the subgroups in alphabetical order to plot them in the same order and colour
        KM_subgroups = OrderedDict(sorted(KM_subgroups.items()))
        
        # Create an empty dictionary to store the KaplanMeierFitter objects
        KMF_object = {}
        logger.info(f"[Subgrouping 4th step] The KM Fitter succesfully calculated the probabilities. \n")
        
        # Create KaplanMeierFitter objects for each subgroup in KM_subgroups
        for label, subset in KM_subgroups.items():
            
            # Remove any rows in the subset with NaN values 
            subset = subset.dropna()

            kmf = KaplanMeierFitter()
            kmf.fit(durations=subset[current_time_column], event_observed=subset[current_event_column])
            KMF_object[label] = kmf

            # Log part of the curve to verify the data was passed correctly
            logger.info(f"[Subgrouping 4th step] Calculated survival function of: {label}")
            logger.info(f"\n {kmf.survival_function_.head(7).to_string()} \n ... \n {kmf.survival_function_.tail(7).to_string()} \n\n")
        
    return KMF_object

###################################################################################################

def save_KM_results(generate_plot_button):

    # Get the required variables from the session state
    KM_analysis_output = st.session_state.get("KM_analysis_output")
    logger = st.session_state.get("logger")
    file_count = st.session_state.get("file_count", 1)

    # File names for Excel and plot
    file_count_str = str(file_count).zfill(2)  # Convert to a 2-digit zero-padded string
    excel_filename = f"KM_results_{file_count_str}.xlsx"
    plot_filename = f"KM_results_{file_count_str}.jpg"

    ###################### Make the plot available for download

    # Get the data from the plot already being displayed
    figure_bytes = st.session_state.get("logged_figure")

    ###################### Make the Excel file available for download

    # Make a new excel if none has been made or the user explicitely wants a new one
    if "logged_excel" not in st.session_state or generate_plot_button:    
        
        # Create a new Excel workbook and remove the default Sheet
        workbook = openpyxl.Workbook()
        workbook.remove(workbook["Sheet"])
        
        # Prepare the data to be processed (single KM object or list of KM objects)
        if isinstance(KM_analysis_output, dict):
            KM_objects_to_process = [{"label": f"KM_Subgroup_{i+1}", "KM_object": KM_object} 
                                    for i, (label, KM_object) in enumerate(KM_analysis_output.items())]
            real_labels = [f"KM_Subgroup_{i+1}: {label}" for i, (label, KM_object) in enumerate(KM_analysis_output.items())]
        else:
            # If KM_analysis_output is a single KM object, add it to the list as a dictionary with a general label
            KM_objects_to_process = [{"label": "KM_Dataset", "KM_object": KM_analysis_output}]
            real_labels = ["KM_Dataset: Whole dataset - No subgroups"]

        # Process all KM curves/objects the same way 
        for index, data in enumerate(KM_objects_to_process):
            
            # Create a sheet per KM object
            sheet = workbook.create_sheet(title=data["label"])

            # Write what the curve/object corresponds to
            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)
            sheet.cell(row=2, column=1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=2, column=1, value=real_labels[index]).font = openpyxl.styles.Font(bold=True, size=16)

            # Get the tables from the KMF object
            event_table = data["KM_object"].event_table
            survival_function = pd.DataFrame({"Time": data["KM_object"].survival_function_.index,
                                            "Survival Probability": np.ravel(data["KM_object"].survival_function_.values)})
            confidence_interval = data["KM_object"].confidence_interval_
            median_survival_time = data["KM_object"].median_survival_time_

            # Write the tables to the Excel sheet
            tables = [event_table, survival_function, confidence_interval, median_survival_time]
            table_names = ["Event Table", "Survival Function", "Confidence Intervals", "Median Survival Time"]
            table_column_numbers = [1, 7, 10, 13]
            
            # Write all tables to the sheet
            for col_index, (table, table_name) in enumerate(zip(tables, table_names)):
                # Define the current column number for the table
                current_column = table_column_numbers[col_index]
        
                # Set the header for the current table
                sheet.cell(row=4, column=current_column, value=table_name).font = openpyxl.styles.Font(bold=True)
        
                if isinstance(table, pd.DataFrame):
                    # If the table is a DataFrame, convert it to a NumPy array
                    rows = table.to_numpy()
        
                    for row_index, row in enumerate(rows):
                        # Write the data from the DataFrame to the Excel sheet
                        for col_offset, value in enumerate(row):
                            sheet.cell(row=row_index + 6, column=current_column + col_offset, value=value)  
                else:
                    # If the table is not a DataFrame, write the single value to the Excel sheet
                    sheet.cell(row=5, column=current_column, value=table)

            ##### Extra worksheet formatting 
            # Merge and center table titles
            sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
            sheet.cell(row=4, column=1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            sheet.merge_cells(start_row=4, start_column=7, end_row=4, end_column=8)
            sheet.cell(row=4, column=7).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            sheet.merge_cells(start_row=4, start_column=10, end_row=4, end_column=11)
            sheet.cell(row=4, column=10).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            
            # Write column titles and center them
            sheet.cell(row=5, column=1, value="Removed")
            sheet.cell(row=5, column=2, value="Observed")
            sheet.cell(row=5, column=3, value="Censored")
            sheet.cell(row=5, column=4, value="Entrance")
            sheet.cell(row=5, column=5, value="At Risk")
            sheet.cell(row=5, column=7, value="Time")
            sheet.cell(row=5, column=8, value="Probability")
            sheet.cell(row=5, column=10, value="Lower Bound")
            sheet.cell(row=5, column=11, value="Upper Bound")
            for cell in ["A5", "B5", "C5", "D5", "E5", "G5", "H5", "J5", "K5"]:
                sheet[cell].alignment = openpyxl.styles.Alignment(horizontal="center")
            
            # Adjust some column widths
            for column, width in zip(["H", "J", "K", "M"], [10, 12, 12, 22]):
                sheet.column_dimensions[column].width = width
            #####
        
        # Save the Excel file to the session state
        excel_bytes = io.BytesIO()
        workbook.save(excel_bytes)
        excel_bytes.seek(0)
        st.session_state["logged_excel"] = excel_bytes
    else:
        # If this rerun does not require making a new excel, use the logged file
        excel_bytes = st.session_state.get("logged_excel")
    
    return figure_bytes, plot_filename, excel_bytes, excel_filename

###################################################################################################
######################################### Flow control ############################################

# Clear the saved df's in case the user wants to upload other files
if restart_button:
    st.session_state.clear()

# Start the processing when the user clicks on the button
if start_button or "flow_control_1" in st.session_state:
    
    # However, proceed only when there have been files uploaded
    if uploaded_files:

        # Create a control for reruns to display the widgets once the user has uploaded
        # files and clicked the Begin button at least once
        st.session_state["flow_control_1"] = True

        # Load the file(s)
        if "df_clinical" not in st.session_state:
            load_input_files(uploaded_files)

        # Process the file(s) to get the information for the main widgets
        if "time_to_event_selection" not in st.session_state and "event_observation_selection" not in st.session_state:
            file_preprocessing()
        
        # Prepare and display the widgets to generate the plots
        widget_preparation()

    else:
        # Display an warning message because we are missing the clinical file
        st.sidebar.warning("Please upload your clinical data file!")

###################################################################################################
